import openpyxl as xl
import pandas as pd
from openpyxl import worksheet
from openpyxl import Workbook
import numpy as np



workbook = xl.load_workbook('Prices.xlsx')
worksheet = workbook.get_sheet_by_name('Sheet1')


columnArray = []
countCol = worksheet.max_column
countRow = worksheet.max_row

#итерация по колонкам и строкам
#создаем из колонок и строк массив
for colunm in worksheet.iter_cols(min_col=2):
    rowArray = []
    for row in range(1, countRow):
        rowArray.append(colunm[row].value)
    columnArray.append(rowArray)
print(columnArray)

#определяем среднее значения массива
stdArray = []
for s in columnArray:
    stdArray.append((np.std(s)) * 2 + np.mean(s))
print(stdArray)

#определить даты и индекс отклонений
arrayondate = []
for col in worksheet.iter_rows(min_row=2):
    array3 = []
    for i in range(1,countCol):
        price = col[i].value
        array3.append(price)
    arrayondate.append(array3)
print(arrayondate)

#сравнить данные в разных колонках
arrayfinish = []
for i in range(len(arrayondate)):
    logic = arrayondate[i] > stdArray
    if logic == True:
        arrayfinish.append(i)
print(arrayfinish)



rownumber = []
for i in arrayfinish:
    i = i+1
    rownumber.append(i)
print(rownumber)

Workbook2 = Workbook(write_only=True)
header = worksheet[1:1]

ws = Workbook2.create_sheet()
ws.append(header)
for rowNum in rownumber:
    row = worksheet[rowNum:rowNum]
    ws.append(row)

Workbook2.save("deviationprice.xlsx")



