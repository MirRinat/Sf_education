
#from src.lessonCode import writeFile, execution as e
import os

import openpyxl as excl

import execution, writeFile
#exlusions = ['оборот','62','62.01','62.02','начальное сальдо','конечное сальдо']
banksSummary = {}

#Есть слово банк, есть запятая, есть пробел. Первый элемент есть счет, который нам интересен
def processBank(contents):
    rowContents = contents.split(' ')
    account = rowContents[0]
    account = account.replace(",",'')
    return(account)

def runAnalysis():
    for file in os.listdir(execution.folder):

        #Первая клетка – компания
        banks = {}
        book = excl.load_workbook(execution.folder+file)
        page = book['TDSheet']
        colA = page['A']
        companyName = colA[0].value
        paymentStartIndex = 0
        #какие строчки рассмартривать
        banks = defineTheRange(banks,colA,companyName,paymentStartIndex)
        # Для выбранных строчек из шага 1:
        for bank in banks:
            monthlyIncome = []
            monthSummary = {}
            startRow = banks[bank][0]
            endRow = banks[bank][1]
            for row in page.iter_rows(min_row=startRow,max_row=endRow):
                # 1.Месяц определяется словом ‘оборот’
                # 2.Каждый месяц
                if row[0].value and 'оборот' in row[0].value.lower():
                    monthlyIncome.append(monthSummary)
                    monthSummary = {}
                # 3.Если не мета-данные, смотрим на дебит
                if row[2].value.lower() not in execution.exlusions:
                    # 5.Делаем сумму по холдинговым компаниям
                    if row[2].value.lower() in execution.holdingsCos:
                        holdingCoName = row[2].value.lower()
                        monthSummary[holdingCoName] = monthSummary.get(holdingCoName,0) + float(row[3].value)
                    # 4.Делаем общую сумму
                    monthSummary[companyName] = monthSummary.get(companyName, 0) + float(row[3].value)

            banksSummary[bank] = monthlyIncome
    # cохранить результат
    writeFile.writeSheet(excl.Workbook(),'Результат',banksSummary)

#Определяем какие строчки рассматривать на основе колнки А
def defineTheRange(banks, colA, companyName, paymentStartIndex):
    for index,cell in enumerate(colA):
        if(cell.value):
            #если есть слово "банк"
            if 'банк ' in cell.value.lower():
                account = processBank(cell.value)
                accountWCoName = account + "____" + companyName
                banks[accountWCoName] = []
            # Мы начинаем смотреть на поступления начиная со строки в которой есть ‘поступлен’ и ‘продаж’
            if 'поступлен' in cell.value.lower() and 'продаж' in cell.value.lower():
                paymentStartIndex = index
            elif paymentStartIndex > 0 and 'оборот' not in cell.value.lower():
                banks[accountWCoName] = [paymentStartIndex,index]
                paymentStartIndex = 0
    return(banks)

runAnalysis()
