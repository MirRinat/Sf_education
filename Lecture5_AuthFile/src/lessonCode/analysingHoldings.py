import os

import openpyxl as excl

import execution, writeFile

IDENTATION = '___'

banksSummary = {}

inputFolder = '../inputData/'

# Есть слово банк, есть запятая, есть пробел. Первый элемент есть счет, который нам интересен
def processBank(contents):
    rowContents = contents.split(' ')
    account = rowContents[0]
    account = account.replace(',','')
    return account

def runAnalysis():
    for file in os.listdir(inputFolder):
        # Первая клетка – компания
        book = excl.load_workbook(inputFolder+file)
        page = book['TDSheet']
        colA = page['A']
        companyName = colA[0].value
        # какие строчки рассмартривать
        entries = {}
        paymentStartIndex = 0
        banks = defineTheRange(entries, colA, companyName, paymentStartIndex)
        sum = 0
        sumHoldings = 0
        for bank in banks:
            monthlyInCome = []
            monthSummary = {}
            firstRow = banks[bank][0]
            lastRow = banks[bank][1]
            for row in page.iter_rows(min_row=firstRow, max_row=lastRow):
                # Для выбранных строчек из шага 1:
                # 1.Месяц определяется словом ‘оборот’
                if row[0].value and 'оборот' in row[0].value.lower():
                # 2.Каждый месяц
                    monthlyInCome.append(monthSummary)
                    monthSummary = {}
                # 3.Если не мета-данные, смотрим на дебит
                if row[2].value.lower() not in execution.exlusions:
                    # 4.Делаем общую сумму
                    monthSummary[companyName] = monthSummary.get(companyName, 0) + int(row[3].value)
                    # 5.Делаем сумму по холдинговым компаниям
                    if row[2].value.lower() in execution.holdingsCos:
                        holdingsCosName = row[2].value.lower()
                        monthSummary[holdingsCosName] = monthSummary.get(holdingsCosName, 0) + int(row[3].value)
            banksSummary[bank] = monthlyInCome
    print(banksSummary)
    writeFile.writeSheet(excl.Workbook(), "Результат", banksSummary)
    # cохранить результат


# Определяем какие строчки рассматривать на основе колнки А
def defineTheRange(banks, colA, companyName, paymentStartIndex):
    # если есть слово 'банк'
    for rowNumber, cell in enumerate(colA):
        if cell.value:
            if 'банк ' in cell.value.lower():
                account =  processBank(cell.value)
                accountWComppany = account + IDENTATION + companyName
                #print(accountWComppany)
            # Мы начинаем смотреть на поступления начиная со строки в которой есть ‘поступлен’ и ‘продаж’
            if 'поступлен' in cell.value.lower() and 'продаж' in cell.value.lower():
                paymentStartIndex = rowNumber
            elif paymentStartIndex > 0 and 'оборот' not in cell.value.lower():
                banks[accountWComppany] = [paymentStartIndex,rowNumber]
                paymentStartIndex = 0
    return banks


runAnalysis()

